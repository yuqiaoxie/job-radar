from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_PATH = PROJECT_ROOT / "data" / "jobs_scored.csv"
OUTPUT_PATH = PROJECT_ROOT / "daily_jobs.xlsx"

COLUMNS = [
    "date_found",
    "title",
    "company",
    "location",
    "source",
    "url",
    "score",
    "matched_keywords",
    "warning_keywords",
    "short_reason",
]


def read_scored_jobs(path: Path = INPUT_PATH) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


def export_excel() -> Path:
    scored_jobs = read_scored_jobs()
    jobs = []
    for job in scored_jobs:
        try:
            score = int(job.get("score", "0"))
            minimum_score = int(job.get("minimum_score_to_export", "60"))
        except ValueError:
            continue
        if job.get("hard_filter_status", "").strip().lower() == "keep" and score > minimum_score:
            jobs.append(job)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Jobs"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for job in jobs:
        sheet.append([job.get(column, "") for column in COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 14,
        "B": 48,
        "C": 24,
        "D": 22,
        "E": 24,
        "F": 54,
        "G": 10,
        "H": 34,
        "I": 28,
        "J": 42,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        row[COLUMNS.index("short_reason")].alignment = Alignment(vertical="top", wrap_text=True)
        score_cell = row[COLUMNS.index("score")]
        try:
            score_cell.value = int(score_cell.value)
        except (TypeError, ValueError):
            pass

    for row_index in range(2, sheet.max_row + 1):
        url_cell = sheet[f"F{row_index}"]
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"

    for column_index in range(1, len(COLUMNS) + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].bestFit = True

    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


def main() -> None:
    path = export_excel()
    print(f"Exported Excel file to {path}")


if __name__ == "__main__":
    main()
